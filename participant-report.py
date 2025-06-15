import re
import logging
import argparse
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
import pdfplumber

dlogging_format = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=dlogging_format)

DEFAULT_SHEET_NAME = '20-TERG-9137_Participation_Atla'
HEADER_ROW = 1
DATA_START_ROW = 20
CLEAR_COLUMNS = ['A', 'B', 'I']
TIME_PATTERN = re.compile(r'^\d+:\d+\.\d+$')


def extract_participant_data(pdf_path):
    participants = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.splitlines():
                    tokens = line.strip().split()
                    if len(tokens) < 6 or not tokens[0].isdigit():
                        continue
                    time_idx = next((i for i, t in enumerate(tokens) if TIME_PATTERN.match(t)), None)
                    if time_idx is None or time_idx < 4:
                        continue
                    number = tokens[1]
                    class_tokens = tokens[time_idx-2:time_idx]
                    name_tokens = tokens[2:time_idx-2]
                    name = ' '.join(name_tokens)
                    cls = ' '.join(class_tokens)
                    participants.append({'Number': number, 'Name': name, 'Class': cls})
    except Exception as e:
        logging.error(f"Error parsing PDF '{pdf_path}': {e}")
        raise

    logging.info(f"Found {len(participants)} participants in PDF.")
    return participants


def update_excel_template(
    pdf_path, template_path, output_path,
    sheet_name=DEFAULT_SHEET_NAME, sheet_password=None
):
    participants = extract_participant_data(pdf_path)

    if not participants:
        logging.error("No participant data extracted. Check PDF format.")
        return

    try:
        book = load_workbook(template_path)
    except Exception as e:
        logging.error(f"Cannot open template '{template_path}': {e}")
        raise

    if sheet_name not in book.sheetnames:
        logging.error(f"Sheet '{sheet_name}' not found. Available: {book.sheetnames}")
        raise KeyError(f"Sheet '{sheet_name}' not found.")

    sheet = book[sheet_name]

    if sheet.protection.sheet:
        logging.info("Unprotecting sheet.")
        sheet.protection = SheetProtection(sheet=False)

    for i in range(max(len(participants), 50)):
        row = DATA_START_ROW + i
        for col in CLEAR_COLUMNS:
            sheet[f"{col}{row}"] = None

    for i, p in enumerate(participants):
        row = DATA_START_ROW + i
        parts = p['Name'].rsplit(' ', 1)
        first = parts[0]
        last = parts[1] if len(parts) == 2 else ''

        sheet[f'A{row}'] = first
        sheet[f'B{row}'] = last
        sheet[f'I{row}'] = p['Class']

    if sheet_password:
        logging.info("Re-protecting sheet.")
        sheet.protection = SheetProtection(password=sheet_password, sheet=True)

    try:
        book.save(output_path)
        logging.info(f"Saved report to '{output_path}'")
    except Exception as e:
        logging.error(f"Error saving to '{output_path}': {e}")
        raise


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Populate participation report')
    parser.add_argument('pdf_input', help='Results PDF path')
    parser.add_argument('excel_template', help='Excel template path')
    parser.add_argument('output_file', help='Output Excel path')
    parser.add_argument('--sheet-name', default=DEFAULT_SHEET_NAME, help='Worksheet name')
    parser.add_argument('--sheet-password', default=None, help='Worksheet password')
    args = parser.parse_args()
    update_excel_template(
        args.pdf_input,
        args.excel_template,
        args.output_file,
        sheet_name=args.sheet_name,
        sheet_password=args.sheet_password
    )